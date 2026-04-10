from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_PATH = Path(r"C:\Users\user\Desktop\boo\記帳表.xlsx")


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(
    left=Side(style="thin", color="B7C9D6"),
    right=Side(style="thin", color="B7C9D6"),
    top=Side(style="thin", color="B7C9D6"),
    bottom=Side(style="thin", color="B7C9D6"),
)


def style_header(ws, row=1):
    for cell in ws[row]:
        if cell.value is None:
            continue
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_table(ws, start_row, end_row, end_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row > 1:
                cell.alignment = Alignment(vertical="center")


def set_widths(ws, widths):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_workbook():
    wb = Workbook()

    ws_txn = wb.active
    ws_txn.title = "交易明細表"
    ws_names = wb.create_sheet("名稱主檔")
    ws_categories = wb.create_sheet("分類主檔")
    ws_accounts = wb.create_sheet("帳戶主檔")
    ws_summary = wb.create_sheet("統計總表")

    build_names_sheet(ws_names)
    build_categories_sheet(ws_categories)
    build_accounts_sheet(ws_accounts)
    build_transactions_sheet(ws_txn)
    build_summary_sheet(ws_summary)
    apply_validations(ws_txn)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2" if ws.title != "統計總表" else "A4"

    return wb


def build_transactions_sheet(ws):
    headers = [
        "日期",
        "單號",
        "交易類型",
        "對象名稱",
        "分類",
        "帳戶",
        "金額",
        "付款狀態",
        "到期日",
        "備註",
    ]
    ws.append(headers)
    style_header(ws)

    example_rows = [
        ["2026-04-01", "TXN-0001", "收入", "王小明", "銷售收入", "銀行", 15000, "已結清", "2026-04-01", "範例資料"],
        ["2026-04-02", "TXN-0002", "支出", "文具行", "辦公用品", "現金", 800, "已結清", "2026-04-02", "範例資料"],
        ["2026-04-03", "TXN-0003", "應收", "客戶A", "專案收入", "銀行", 12000, "未結清", "2026-04-30", "待收款"],
        ["2026-04-04", "TXN-0004", "應付", "供應商B", "進貨", "信用卡", 5000, "部分付款", "2026-04-20", "月底結清"],
    ]
    for row in example_rows:
        ws.append(row)

    for row in range(2, 2002):
        ws[f"K{row}"] = f'=IF(OR(C{row}="應收",C{row}="借出"),IF(H{row}<>"已結清",G{row},0),0)'
        ws[f"L{row}"] = f'=IF(OR(C{row}="應付",C{row}="借入"),IF(H{row}<>"已結清",G{row},0),0)'

    ws["K1"] = "應收未收"
    ws["L1"] = "應付未付"
    style_header(ws)
    style_table(ws, 1, 2001, 12)
    set_widths(
        ws,
        {
            1: 14,
            2: 14,
            3: 14,
            4: 18,
            5: 18,
            6: 14,
            7: 14,
            8: 14,
            9: 14,
            10: 24,
            11: 14,
            12: 14,
        },
    )


def build_names_sheet(ws):
    headers = ["名稱", "類型", "備註"]
    ws.append(headers)
    style_header(ws)
    rows = [
        ["王小明", "人名", ""],
        ["文具行", "廠商", ""],
        ["客戶A", "客戶", ""],
        ["供應商B", "廠商", ""],
        ["銀行", "帳戶", ""],
        ["現金", "帳戶", ""],
        ["信用卡", "帳戶", ""],
        ["銷售收入", "分類", ""],
        ["辦公用品", "分類", ""],
    ]
    for row in rows:
        ws.append(row)
    style_table(ws, 1, max(ws.max_row, 200), 3)
    set_widths(ws, {1: 20, 2: 14, 3: 24})


def build_categories_sheet(ws):
    headers = ["分類名稱", "分類群組", "備註"]
    ws.append(headers)
    style_header(ws)
    rows = [
        ["餐飲", "支出", ""],
        ["交通", "支出", ""],
        ["進貨", "支出", ""],
        ["薪資", "支出", ""],
        ["借款", "資金往來", ""],
        ["還款", "資金往來", ""],
        ["銷售收入", "收入", ""],
        ["專案收入", "收入", ""],
        ["辦公用品", "支出", ""],
    ]
    for row in rows:
        ws.append(row)
    style_table(ws, 1, max(ws.max_row, 200), 3)
    set_widths(ws, {1: 20, 2: 14, 3: 24})


def build_accounts_sheet(ws):
    headers = ["帳戶名稱", "帳戶類型", "備註"]
    ws.append(headers)
    style_header(ws)
    rows = [
        ["現金", "現金", ""],
        ["銀行", "銀行", ""],
        ["電子支付", "電子支付", ""],
        ["信用卡", "信用卡", ""],
    ]
    for row in rows:
        ws.append(row)
    style_table(ws, 1, max(ws.max_row, 200), 3)
    set_widths(ws, {1: 20, 2: 16, 3: 24})


def build_summary_sheet(ws):
    ws["A1"] = "統計總表"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "項目"
    ws["B3"] = "數值"
    for cell in ("A3", "B3"):
        ws[cell].fill = HEADER_FILL
        ws[cell].font = HEADER_FONT
        ws[cell].alignment = Alignment(horizontal="center")
        ws[cell].border = THIN_BORDER

    metrics = [
        ("總收入", '=SUMIF(交易明細表!C:C,"收入",交易明細表!G:G)+SUMIF(交易明細表!C:C,"收款",交易明細表!G:G)'),
        ("總支出", '=SUMIF(交易明細表!C:C,"支出",交易明細表!G:G)+SUMIF(交易明細表!C:C,"付款",交易明細表!G:G)'),
        ("應收未收", '=SUM(交易明細表!K:K)'),
        ("應付未付", '=SUM(交易明細表!L:L)'),
    ]
    row = 4
    for label, formula in metrics:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = formula
        row += 1

    ws["D3"] = "分類名稱"
    ws["E3"] = "合計"
    ws["G3"] = "對象名稱"
    ws["H3"] = "往來餘額"
    for cell in ("D3", "E3", "G3", "H3"):
        ws[cell].fill = HEADER_FILL
        ws[cell].font = HEADER_FONT
        ws[cell].alignment = Alignment(horizontal="center")
        ws[cell].border = THIN_BORDER

    for idx in range(2, 102):
        summary_row = idx + 2
        ws[f"D{summary_row}"] = f'=IF(分類主檔!A{idx}="","",分類主檔!A{idx})'
        ws[f"E{summary_row}"] = (
            f'=IF(D{summary_row}="","",SUMIF(交易明細表!E:E,D{summary_row},交易明細表!G:G))'
        )
        ws[f"G{summary_row}"] = f'=IF(名稱主檔!A{idx}="","",名稱主檔!A{idx})'
        ws[f"H{summary_row}"] = (
            f'=IF(G{summary_row}="","",'
            f'SUMIFS(交易明細表!G:G,交易明細表!D:D,G{summary_row},交易明細表!C:C,"收入")'
            f'+SUMIFS(交易明細表!G:G,交易明細表!D:D,G{summary_row},交易明細表!C:C,"收款")'
            f'+SUMIFS(交易明細表!G:G,交易明細表!D:D,G{summary_row},交易明細表!C:C,"借入")'
            f'-SUMIFS(交易明細表!G:G,交易明細表!D:D,G{summary_row},交易明細表!C:C,"支出")'
            f'-SUMIFS(交易明細表!G:G,交易明細表!D:D,G{summary_row},交易明細表!C:C,"付款")'
            f'-SUMIFS(交易明細表!G:G,交易明細表!D:D,G{summary_row},交易明細表!C:C,"借出"))'
        )

    style_table(ws, 3, 104, 8)
    set_widths(ws, {1: 18, 2: 16, 4: 18, 5: 16, 7: 18, 8: 16})


def apply_validations(ws):
    type_validation = DataValidation(
        type="list",
        formula1='"收入,支出,應收,收款,應付,付款,借出,借入"',
        allow_blank=True,
    )
    status_validation = DataValidation(
        type="list",
        formula1='"已結清,未結清,部分付款"',
        allow_blank=True,
    )
    name_validation = DataValidation(
        type="list",
        formula1="=名稱主檔!$A$2:$A$200",
        allow_blank=True,
    )
    category_validation = DataValidation(
        type="list",
        formula1="=分類主檔!$A$2:$A$200",
        allow_blank=True,
    )
    account_validation = DataValidation(
        type="list",
        formula1="=帳戶主檔!$A$2:$A$200",
        allow_blank=True,
    )

    for validation in [type_validation, status_validation, name_validation, category_validation, account_validation]:
        ws.add_data_validation(validation)

    type_validation.add("C2:C2001")
    name_validation.add("D2:D2001")
    category_validation.add("E2:E2001")
    account_validation.add("F2:F2001")
    status_validation.add("H2:H2001")


def main():
    wb = create_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
