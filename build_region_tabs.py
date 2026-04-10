from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\user\Documents\New project\報表.xlsx")
OUTPUT = Path(r"C:\Users\user\Documents\New project\報表_自動月報.xlsm")
REGIONS = ["台北", "桃園", "新竹", "台中", "台南", "高雄"]


def normalize_number(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return value
    try:
        return float(value)
    except Exception:
        return 0


def get_rows(ws_formula, ws_value):
    rows = []
    for row in range(2, ws_formula.max_row + 1):
        month = ws_formula.cell(row, 1).value
        region = ws_formula.cell(row, 15).value
        if month is None or region is None:
            continue

        unit_price = normalize_number(ws_value.cell(row, 11).value)
        qty = normalize_number(ws_value.cell(row, 9).value)
        untaxed = ws_value.cell(row, 12).value
        tax = ws_value.cell(row, 13).value
        subtotal = ws_value.cell(row, 14).value
        untaxed = normalize_number(untaxed) if untaxed is not None else unit_price * qty
        tax = normalize_number(tax)
        subtotal = normalize_number(subtotal) if subtotal is not None else untaxed + tax

        rows.append(
            {
                "month": month,
                "invoice_type": ws_formula.cell(row, 5).value,
                "region": str(region).strip(),
                "customer": ws_formula.cell(row, 6).value or "",
                "company": ws_formula.cell(row, 18).value or "",
                "report_name": ws_formula.cell(row, 7).value or "",
                "item": ws_formula.cell(row, 8).value or "",
                "untaxed": untaxed,
                "tax": tax,
                "subtotal": subtotal,
                "invoice_no": ws_formula.cell(row, 3).value or "",
                "note": ws_formula.cell(row, 16).value or "",
            }
        )
    return rows


def write_region_sheet(ws, rows, region):
    ws.delete_rows(1, ws.max_row)
    current_row = 1

    if not rows:
        ws.cell(current_row, 1, "區域")
        ws.cell(current_row, 2, region)
        ws.cell(current_row + 1, 1, "狀態")
        ws.cell(current_row + 1, 2, "目前無資料")
        return

    buckets = defaultdict(list)
    order = []
    for item in rows:
        key = (item["month"], item["invoice_type"])
        if key not in buckets:
            order.append(key)
        buckets[key].append(item)

    for month, invoice_type in order:
        group_rows = buckets[(month, invoice_type)]
        ws.cell(current_row, 1, "請款月份")
        ws.cell(current_row, 2, month)
        ws.cell(current_row + 1, 1, "發票別")
        ws.cell(current_row + 1, 2, invoice_type)
        ws.cell(current_row + 2, 1, "區域")
        ws.cell(current_row + 2, 2, region)

        headers = ["客戶別", "公司抬頭", "報表名稱", "項目", "未稅額 ", "稅金 ", "小計 ", "發票號碼", "備註"]
        for idx, header in enumerate(headers, start=1):
            ws.cell(current_row + 3, idx, header)

        detail_row = current_row + 4
        customer_buckets = defaultdict(list)
        customer_order = []
        for item in group_rows:
            customer = item["customer"]
            if customer not in customer_buckets:
                customer_order.append(customer)
            customer_buckets[customer].append(item)

        for customer in customer_order:
            customer_rows = customer_buckets[customer]
            for item in customer_rows:
                ws.cell(detail_row, 1, item["customer"])
                ws.cell(detail_row, 2, item["company"])
                ws.cell(detail_row, 3, item["report_name"])
                ws.cell(detail_row, 4, item["item"])
                ws.cell(detail_row, 5, item["untaxed"])
                ws.cell(detail_row, 6, item["tax"])
                ws.cell(detail_row, 7, item["subtotal"])
                ws.cell(detail_row, 8, item["invoice_no"])
                ws.cell(detail_row, 9, item["note"])
                detail_row += 1

            ws.cell(detail_row, 1, f"{customer} 合計")
            ws.cell(detail_row, 5, f"=SUM(E{detail_row - len(customer_rows)}:E{detail_row - 1})")
            ws.cell(detail_row, 6, f"=SUM(F{detail_row - len(customer_rows)}:F{detail_row - 1})")
            ws.cell(detail_row, 7, f"=SUM(G{detail_row - len(customer_rows)}:G{detail_row - 1})")
            detail_row += 2

        current_row = detail_row

    widths = [14, 18, 18, 28, 12, 12, 12, 14, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def main():
    wb_formula = load_workbook(SOURCE, data_only=False)
    wb_value = load_workbook(SOURCE, data_only=True)
    ws_formula = wb_formula["總表"]
    ws_value = wb_value["總表"]

    all_rows = get_rows(ws_formula, ws_value)

    for region in REGIONS:
        sheet_name = f"{region}月報"
        if sheet_name in wb_formula.sheetnames:
            ws = wb_formula[sheet_name]
        else:
            ws = wb_formula.create_sheet(sheet_name)

        region_rows = [r for r in all_rows if r["region"] == region]
        write_region_sheet(ws, region_rows, region)

    wb_formula.save(OUTPUT)
    print(str(OUTPUT))
    for region in REGIONS:
        count = sum(1 for r in all_rows if r["region"] == region)
        print(f"{region}:{count}")


if __name__ == "__main__":
    main()
